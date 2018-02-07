#!/usr/bin/python3.3
#_*_coding:utf-8_*_
#
# 文件作用：电子邮件工具类
#
# 创建时间：2015-07-06
#
#

from openpyxl import Workbook
from mongoengine import ReferenceField
from common.models import *
from mongoengine import ListField
from mongoengine.base.datastructures import BaseList
import datetime
from common.tools_legoo import DocumentTools, DocumentFolderType, ImageTools, ImageFolderType
try:
    from common.models import SiteSettings
except ImportError:
    from common.models import BaseSettings as SiteSettings
from common.models import DException
import hashlib
from django.conf import settings
import traceback
from uuid import uuid1
import os
#添加时间判断部分
from datetime import timedelta, date


class ExcelExportTools(object):

    def get_data(self, value, field_key, field_name):
        if value is None:
            return ''
        elif isinstance(value, str):
            count = value.find('/')
            if count > 2:
                return settings.SERVER_URL + '/static/' + value
            if value == 'platform':
                return '平台上传'
            if value == 'input':
                return '系统录入'
            if value == 'submit':
                return '公众号投保'
            if field_key == 'startSiteName' or field_key== "targetSiteName" :
                address_number = value
                try:
                    print(address_number)
                    name_detail = address_number.split(" ")
                    count = len(name_detail)
                    address_detail=""
                    for i in range(count):
                        prov_code = name_detail[i]
                        if prov_code:
                            prov_name = CargoArea.objects(code=prov_code).first().name
                            address_detail = address_detail + prov_name+" "
                        else:
                            continue
                    return address_detail.strip()
                except:
                    message = address_number+"未找到订单对应编码的地址名称,请联系运至宝管理员"
                    return message
            return value
        elif isinstance(value, int):
            if field_key == 'insurance_price' or field_key == 'pay_price':
                value = round(value/100.0, 2)
            return value
        elif isinstance(value, dict):
            return str(value)
        elif isinstance(value, datetime.datetime):
            return str(value.strftime("%Y-%m-%d"))
        elif isinstance(value, ReferenceField):
            return str(value.id)
        elif isinstance(value, InsuranceCompany):
            return str(value.simple_name)
        elif isinstance(value, InsuranceProducts):
            return str(value.name)
        elif isinstance(value, Client):
            if field_name == '投保人':
                if value.company_name:
                    return str(value.company_name)
                else:
                    return str(value.name)
            elif field_name == '推荐人':
                client = Client.objects(user=value.referee).first()
                if client:
                    return str(client.profile.phone)
                else:
                    return '无推荐人'
        elif isinstance(value, float):
            value = value * 100000 / 100000
            return value
        elif isinstance(value, BaseList):
            if field_key == 'batch_image_list':
                batch_image_list = ''
                for batch_image in value:
                    batch_image_list += settings.SERVER_URL + '/static/' + batch_image + '   '
                return batch_image_list
            elif field_key == 'batch_list':
                bacth_str = ''
                for bacth in value:
                    bacth_str += bacth.transport_id + ' ' + bacth.startSiteName + ' 至 ' + bacth.targetSiteName + ' ' + bacth.commodityName + ' ' + str(bacth.commodityCases) + ' '
                return str(bacth_str)
        else:
            print('unknown type {0} of value {1}'.format(value, type(value)))
            return str(value)

    def export(self, query_set, field_tuple):
        print("开始写入")
        wb = Workbook()

        file_time = datetime.datetime.now().strftime("%Y%m%d%H%I%S")
        file_name = file_time + ".xlsx"
        batch = 'order/list'
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        # 基本目录,若基本目录不存在则需要创建基本目录
        base_dir = "{0}/static/{1}".format(settings.BASE_DIR, batch)
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        # 存放文件的目录，若目录不存在则需要创建目录
        file_dir = "{0}/static/{1}/{2}".format(settings.BASE_DIR, batch, date_str)
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
        # 文件的相对路径，数据库中存放的路径
        file_res_path = "{0}/{1}/{2}".format(batch, date_str, file_name)
        # 文件的绝对路径，文件系统中文件的路径
        file_abs_path = "{0}/static/{1}".format(settings.BASE_DIR, file_res_path)


        print("创建表")
        sheet = wb.create_sheet("订单列表", 0)
        print("写入表头")
        for position, item in enumerate(field_tuple):
            key, name = item
            sheet.cell(row=1, column=position+1).value = name
        print("开始写入数据")
        for row_position, item in enumerate(query_set):
            print("写入第{0}个数据".format(row_position+1))
            for column_position, field in enumerate(field_tuple):
                field_key, field_name = field
                if field_key == 'startSiteName'  or   field_key== "targetSiteName" :
                    test_city_detail = getattr(item, field_key)
                    if   test_city_detail :
                        sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data(getattr(item, field_key), field_key, field_name)
                    else:
                        if field_key == 'startSiteName' :
                            sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data(getattr(item, 'car_startSiteName'), 'car_startSiteName', field_name)
                        elif field_key== "targetSiteName" :
                            sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data(getattr(item, 'car_targetSiteName'), 'car_targetSiteName', field_name)
                else:
                    sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data(getattr(item, field_key), field_key, field_name)
                # sheet.cell(row=1, column=position+1).value = name
        wb.save(file_abs_path)
        print("写入数据成功！")
        return file_res_path
#机动车辆保险导出保单  
    def get_data1(self, value, field_key, field_name):
        field_key=field_key
        field_name =field_name
        if value is None:
            return ''
        elif isinstance(value, ReferenceField):
            return str(value.id)
        elif isinstance(value, datetime.datetime):
            return str(value.strftime("%Y-%m-%d"))
        elif isinstance(value, Ordering):
            return str(value.paper_id)
        elif isinstance(value, InquiryInfo):
            return str(value.paper_id)
        elif isinstance(value, Client):
            return str(value.profile.phone)
        elif isinstance(value, InsuranceCompany):
            return str(value.name)
        elif isinstance(value, Intermediary):
            detai=str(value.intermediary_name) +'报案电话：'+ str(value.intermediary_phone)#+'利润点：'+ str(value.intermediary_profit_point)
            return str(detai)
        if field_key == 'state'  :
            order_type_detail = value
            order_type_name=''
            for type_detai in InquiryInfo.ORDER_TYPE:
                if type_detai[0]==order_type_detail:
                    order_type_name = type_detai[1]
            return str(order_type_name)
        elif field_key == 'car_type'  :
            order_type_detail = value
            order_type_name=''
            for type_detai in InquiryInfo.CAR_TYPE:
                if type_detai[0]==order_type_detail:
                    order_type_name = type_detai[1]
            return str(order_type_name)
        elif field_key == 'price'   or  field_key == 'intermediary_price' or  field_key == 'pay_price'  :
            value=float(value)/100
            return str(value)
        else:
            print('unknown type {0} of value {1}'.format(value, type(value)))
            return str(value)
        
        
    def get_data2(self, value, field_key, field_name):
        field_key=field_key
        field_name =field_name
        special_list =[ 'quoted_profit_point' , 'liability_price' ,'commercial_price' ,'vehicle_vessel_price']
        if value is None:
            return ''
        elif isinstance(value, InquiryInfo):
            if field_key == 'profit'  :
                profit=(float(value.pay_price) - float(value.intermediary_price))/100
                return str(profit)
            elif field_key == 'record_clerk'  :
                return str('admin')
            elif field_key == 'referee_people'  :
                client = value.client
                #查找推荐人信息
                try:
                    referee = Client.objects(user=client.referee).first() 
                    if not referee:
                        return str('无')
                    else:
                        referee_detail1=''
                        if referee.name:
                            referee_detail1 ="推荐人姓名："+referee.name
                        elif referee.company_name:
                            referee_detail1 ="推荐人单位名称："+referee.company_name
                        referee_detail = referee_detail1+"推荐人手机号："+referee.profile.phone
                        return str(referee_detail)
                except:
                    return str('无')
            elif field_key in special_list:
                try:
                    intermediary_price_set = IntermediaryPrice.objects(order=value,company=value.company,insurance_intermediary=value.insurance_intermediary,order_price_add_profit=value.price).first()
                    if not intermediary_price_set:
                        return '未找到对应数据。'
                    else:
                        return_detail=''
                        if field_key == 'quoted_profit_point'  :
                            return_detail=str(intermediary_price_set.intermediary_profit_point)+'%'
                        elif field_key == 'liability_price'  :
                            return_detail=str(float(intermediary_price_set.liability_price)/100)+'元'
                        elif field_key == 'vehicle_vessel_price'  :
                            return_detail=str(float(intermediary_price_set.vehicle_vessel_price)/100)+'元'
                        elif field_key == 'commercial_price'  :
                            return_detail=str(float(intermediary_price_set.commercial_price)/100)+'元'
                        return str(return_detail)
                except:
                    return  '未找到对应数据'
            elif field_key == 'liability_expectEndTime':
                if value.liability_expectStartTime:
                    liability_expectEndTime = value.liability_expectStartTime + timedelta(days =365)
                    return str(liability_expectEndTime.strftime("%Y-%m-%d"))
                else:
                    return ''
            elif field_key == 'commercial_expectEndTime':
                if value.commercial_expectStartTime:
                    commercial_expectEndTime = value.commercial_expectStartTime + timedelta(days =365)
                    return str(commercial_expectEndTime.strftime("%Y-%m-%d"))
                else:
                    return ''
            else:
                return ''
        else:
            return ''
        
    def export1(self, query_set, field_tuple):
        print("开始写入")
        special_list =['referee_people' , 'profit' , 'quoted_profit_point' , 'liability_price' ,'commercial_price' ,'vehicle_vessel_price','commercial_expectEndTime' ,'liability_expectEndTime','record_clerk']
        wb = Workbook()

        file_time = datetime.datetime.now().strftime("%Y%m%d%H%I%S")
        file_name = file_time + ".xlsx"
        batch = 'order/list'
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        # 基本目录,若基本目录不存在则需要创建基本目录
        base_dir = "{0}/static/{1}".format(settings.BASE_DIR, batch)
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        # 存放文件的目录，若目录不存在则需要创建目录
        file_dir = "{0}/static/{1}/{2}".format(settings.BASE_DIR, batch, date_str)
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
        # 文件的相对路径，数据库中存放的路径
        file_res_path = "{0}/{1}/{2}".format(batch, date_str, file_name)
        # 文件的绝对路径，文件系统中文件的路径
        file_abs_path = "{0}/static/{1}".format(settings.BASE_DIR, file_res_path)


        print("创建表")
        sheet = wb.create_sheet("订单列表", 0)
        print("写入表头")
        for position, item in enumerate(field_tuple):
            key, name = item
            sheet.cell(row=1, column=position+1).value = name
        print("开始写入数据")
        for row_position, item in enumerate(query_set):
            print("写入第{0}个数据".format(row_position+1))
            for column_position, field in enumerate(field_tuple):
                field_key, field_name = field
                if field_key not in special_list:
                    sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data1(getattr(item, field_key), field_key, field_name)
                else:
                    sheet.cell(row=row_position +2, column=column_position +1).value = self.get_data2(item, field_key, field_name)
                # sheet.cell(row=1, column=position+1).value = name
        wb.save(file_abs_path)
        print("写入数据成功！")
        return file_res_path
    


